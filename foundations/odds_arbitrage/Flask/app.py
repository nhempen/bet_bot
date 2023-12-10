from flask import Flask, render_template, request, make_response
import subprocess
from openpyxl import load_workbook

# (A2) FLASK SETTINGS + INIT
HOST_NAME = "0.0.0.0"
HOST_PORT = 8080
app = Flask(__name__)
app.debug = True



@app.route('/refresh_odds', methods=['POST'])
def fetch_odds_data():
    cmd = ["python3", "collect_odds.py"]
    stdout = subprocess.check_output(cmd).decode()
    book = load_workbook("latest_odds.xlsx")
    sheet = book.active
    return render_template("index.html", sheet=sheet)


# (B) DEMO - READ EXCEL & GENERATE HTML TABLE
@app.route("/")
def index():
    cmd = ["python3", "collect_odds.py"]
    stdout = subprocess.check_output(cmd).decode()
    # (B1) OPEN EXCEL FILE + WORKSHEET
    book = load_workbook("latest_odds.xlsx")
    sheet = book.active

    # (B2) PASS INTO HTML TEMPLATE
    return render_template("index.html", sheet=sheet)

# (C) START
if __name__ == "__main__":
    app.run(HOST_NAME, HOST_PORT)